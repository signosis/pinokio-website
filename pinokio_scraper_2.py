# Create a merged, non-interactive GitHub Actions–friendly script
merged_script = r'''#!/usr/bin/env python3
import os
import time
import base64
import json
import sqlite3
from datetime import datetime, timezone
import csv
import re
from typing import Optional, Dict, Any, List

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ================================
# CONFIG (env-driven, non-interactive)
# ================================
PINOKIO_ORG = os.getenv("PINOKIO_ORG", "pinokiofactory")
API_URL = "https://api.github.com"

# Prefer GH_TOKEN (GitHub Actions convention), fall back to GITHUB_TOKEN
GITHUB_TOKEN = os.getenv("GH_TOKEN") or os.getenv("GITHUB_TOKEN") or ""
HEADERS = {"Accept": "application/vnd.github.v3+json"}
if GITHUB_TOKEN:
    HEADERS["Authorization"] = f"token {GITHUB_TOKEN}"

# Output paths (these are safe for Actions artifacts or Pages)
DB_FILE = os.getenv("DB_FILE", "pinokio.db")
CSV_FILE = os.getenv("CSV_FILE", "pinokio_repos.csv")
XLSX_FILE = os.getenv("XLSX_FILE", "pinokio_repos.xlsx")
JSON_FILE = os.getenv("JSON_FILE", "docs/data.json")

# Behavior flags
FORCE_REFRESH = os.getenv("REFRESH", "0") in ("1", "true", "True", "YES", "yes")

# ================================
# DB SETUP (merged schema)
# ================================
def init_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS repos (
            id INTEGER PRIMARY KEY,
            name TEXT UNIQUE,
            description TEXT,
            html_url TEXT,
            created_at TEXT,
            updated_at TEXT,
            pushed_at TEXT,
            open_issues INTEGER,
            upstream_name TEXT,
            upstream_url TEXT,
            upstream_created_at TEXT,
            upstream_updated_at TEXT,
            upstream_open_issues INTEGER,
            last_checked TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS api_usage (
            last_run TEXT,
            calls_used INTEGER,
            calls_remaining INTEGER
        )
    """)
    conn.commit()
    return conn

# ================================
# API helpers with rate handling
# ================================
_api_calls_used = 0

def _github_get(url: str, headers: Dict[str, str]) -> requests.Response:
    """GET request counting API calls and handling rate limit politely (no prompts)."""
    global _api_calls_used
    while True:
        r = requests.get(url, headers=headers)
        _api_calls_used += 1
        # Handle rate limiting
        if r.status_code == 403 and "X-RateLimit-Remaining" in r.headers:
            if r.headers.get("X-RateLimit-Remaining") == "0":
                reset = int(r.headers.get("X-RateLimit-Reset", time.time() + 60))
                wait = max(reset - int(time.time()), 1)
                print(f"[rate-limit] Remaining=0. Sleeping {wait}s until reset...")
                time.sleep(wait)
                continue
        r.raise_for_status()
        return r

def _check_rate_limit() -> Dict[str, Any]:
    r = requests.get(f"{API_URL}/rate_limit", headers=HEADERS)
    r.raise_for_status()
    data = r.json()
    core = data.get("rate", {})
    return {
        "limit": core.get("limit", 0),
        "remaining": core.get("remaining", 0),
        "reset": core.get("reset", int(time.time())),
    }

# ================================
# Repo helpers
# ================================
def list_org_repos(org: str) -> List[Dict[str, Any]]:
    url = f"{API_URL}/orgs/{org}/repos?per_page=100"
    repos = []
    while url:
        r = _github_get(url, HEADERS)
        repos.extend(r.json())
        url = r.links.get("next", {}).get("url")
    return repos

def get_repo_info(owner: str, repo: str) -> Dict[str, Any]:
    url = f"{API_URL}/repos/{owner}/{repo}"
    r = _github_get(url, HEADERS)
    data = r.json()
    return {
        "full_name": data["full_name"],
        "description": data.get("description") or "",
        "created_at": data["created_at"],
        "updated_at": data["updated_at"],
        "pushed_at": data.get("pushed_at"),
        "open_issues": data.get("open_issues_count", 0),
        "html_url": data["html_url"],
    }

def _decode_content_json(org: str, repo: str, path: str) -> Optional[Dict[str, Any]]:
    """Fetch a JSON file via the contents API and decode base64 content."""
    try:
        url = f"{API_URL}/repos/{org}/{repo}/contents/{path}"
        r = _github_get(url, HEADERS)
        content = r.json()
        text = base64.b64decode(content["content"]).decode("utf-8", errors="replace")
        return json.loads(text)
    except Exception:
        return None

def _fetch_readme_text(org: str, repo: str) -> Optional[str]:
    """Try raw README first (cheap), fallback to contents API."""
    raw_url = f"https://raw.githubusercontent.com/{org}/{repo}/main/README.md"
    r = requests.get(raw_url, headers={"Accept": "text/plain"})
    if r.status_code == 200 and r.text:
        return r.text
    # fallback: use contents API (base64)
    try:
        url = f"{API_URL}/repos/{org}/{repo}/contents/README.md"
        r2 = _github_get(url, HEADERS)
        content = r2.json()
        return base64.b64decode(content["content"]).decode("utf-8", errors="replace")
    except Exception:
        return None

def _extract_github_url_from_json(data: Any) -> Optional[str]:
    """Walk any JSON to find a string containing github.com/owner/repo."""
    if isinstance(data, dict):
        for v in data.values():
            u = _extract_github_url_from_json(v)
            if u:
                return u
    elif isinstance(data, list):
        for v in data:
            u = _extract_github_url_from_json(v)
            if u:
                return u
    elif isinstance(data, str):
        if "github.com" in data:
            m = re.search(r"https://github\.com/[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+", data)
            if m:
                return m.group(0)
    return None

def _extract_github_url_from_readme(text: str) -> Optional[str]:
    m = re.search(r"https://github\.com/[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+", text or "")
    return m.group(0) if m else None

def find_upstream(org: str, repo: str) -> Optional[Dict[str, Any]]:
    """
    Look for upstream via pinokio.json, install.json, or README.md.
    Return upstream repo info dict or None.
    """
    # Try JSON recipes
    for fname in ("pinokio.json", "install.json"):
        data = _decode_content_json(org, repo, fname)
        if data:
            u = _extract_github_url_from_json(data)
            if u:
                try:
                    owner, rname = u.rstrip("/").split("/")[-2:]
                    return get_repo_info(owner, rname)
                except Exception:
                    pass
    # Fallback: README
    text = _fetch_readme_text(org, repo)
    if text:
        u = _extract_github_url_from_readme(text)
        if u:
            try:
                owner, rname = u.rstrip("/").split("/")[-2:]
                return get_repo_info(owner, rname)
            except Exception:
                pass
    return None

# ================================
# Export helpers
# ================================
def export_csv(rows: List[Dict[str, Any]], path: str) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ CSV saved: {path}")

def export_xlsx(rows: List[Dict[str, Any]], path: str) -> None:
    if not rows:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Pinokio Repos"
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    # Hyperlinks
    for rr in range(2, len(rows) + 2):
        for cc, header in enumerate(headers, start=1):
            val = ws.cell(row=rr, column=cc).value
            if isinstance(val, str) and header.endswith("url") and val.startswith("http"):
                ws.cell(row=rr, column=cc).hyperlink = val
                ws.cell(row=rr, column=cc).style = "Hyperlink"
    # Autofit (approximate)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
    wb.save(path)
    print(f"✅ XLSX saved: {path}")

def export_json(rows: List[Dict[str, Any]], path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    print(f"✅ JSON saved: {path}")

# ================================
# Main
# ================================
def main() -> None:
    rl = _check_rate_limit()
    reset_dt = datetime.fromtimestamp(rl["reset"], tz=timezone.utc)
    print(f"[rate-limit] limit={rl['limit']} remaining={rl['remaining']} reset={reset_dt.isoformat()}")

    conn = init_db(DB_FILE)
    cur = conn.cursor()

    repos = list_org_repos(PINOKIO_ORG)
    rows_out: List[Dict[str, Any]] = []

    for repo in repos:
        name = repo["name"]
        description = repo.get("description", "") or ""
        html_url = repo["html_url"]
        created_at = repo["created_at"]
        updated_at = repo["updated_at"]
        pushed_at = repo.get("pushed_at")
        open_issues = repo.get("open_issues_count", 0)

        # Cache policy: update if (FORCE_REFRESH) or (db row missing) or (updated_at changed)
        cur.execute("SELECT updated_at FROM repos WHERE name=?", (name,))
        row = cur.fetchone()
        should_refresh = FORCE_REFRESH or (row is None) or (row and row[0] != updated_at)

        upstream_name = upstream_url = up_created = up_updated = None
        up_issues = None

        if should_refresh:
            upstream = find_upstream(PINOKIO_ORG, name)
            if upstream:
                upstream_name = upstream["full_name"]
                upstream_url = upstream["html_url"]
                up_created = upstream["created_at"]
                up_updated = upstream["updated_at"]
                up_issues = upstream["open_issues"]

            cur.execute("""
                INSERT OR REPLACE INTO repos
                (id, name, description, html_url, created_at, updated_at, pushed_at, open_issues,
                 upstream_name, upstream_url, upstream_created_at, upstream_updated_at, upstream_open_issues,
                 last_checked)
                VALUES (
                    (SELECT id FROM repos WHERE name=?),
                    ?,?,?,?,?,?,?,
                    ?,?,?,?,?,
                    ?
                )
            """, (
                name,
                name, description, html_url, created_at, updated_at, pushed_at, open_issues,
                upstream_name, upstream_url, up_created, up_updated, up_issues,
                datetime.now(timezone.utc).isoformat()
            ))
            conn.commit()

        # Emit row (from freshest known state)
        cur.execute("SELECT name, description, html_url, created_at, updated_at, pushed_at, open_issues, upstream_name, upstream_url, upstream_created_at, upstream_updated_at, upstream_open_issues FROM repos WHERE name=?", (name,))
        dbrow = cur.fetchone()
        if dbrow:
            rows_out.append({
                "name": dbrow[0],
                "description": dbrow[1],
                "html_url": dbrow[2],
                "created_at": dbrow[3],
                "updated_at": dbrow[4],
                "pushed_at": dbrow[5],
                "open_issues": dbrow[6],
                "upstream_name": dbrow[7] or "",
                "upstream_url": dbrow[8] or "",
                "upstream_created_at": dbrow[9] or "",
                "upstream_updated_at": dbrow[10] or "",
                "upstream_open_issues": dbrow[11] if dbrow[11] is not None else "",
            })

    # Save API usage snapshot
    try:
        cur.execute("INSERT INTO api_usage VALUES (?, ?, ?)", (
            datetime.now(timezone.utc).isoformat(),
            _api_calls_used,
            max(0, rl["remaining"] - _api_calls_used),
        ))
        conn.commit()
    except Exception:
        pass

    conn.close()

    # Exports
    rows_out.sort(key=lambda r: (r["upstream_name"], r["name"]))
    export_csv(rows_out, CSV_FILE)
    export_xlsx(rows_out, XLSX_FILE)
    export_json(rows_out, JSON_FILE)

    print(f"[done] repos={len(rows_out)} api_calls={_api_calls_used}")

if __name__ == "__main__":
    main()
'''
# Write to file
with open('/mnt/data/pinokio_scraper_merged.py', 'w', encoding='utf-8') as f:
    f.write(merged_script)

print("Created /mnt/data/pinokio_scraper_merged.py")
